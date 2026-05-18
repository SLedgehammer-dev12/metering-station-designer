"""Excel compliance report for geometric inspection results."""

import io
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_compliance_report(report, lang: str = "tr") -> io.BytesIO:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl gereklidir")

    wb = Workbook()

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    pass_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    fail_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    cond_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def status_fill(status):
        if status == "PASS":
            return pass_fill
        elif status == "FAIL":
            return fail_fill
        elif status == "CONDITIONAL":
            return cond_fill
        return None

    # Sheet 1: Overview
    ws1 = wb.active
    ws1.title = "Ozet" if lang == "tr" else "Overview"
    ws1.cell(row=1, column=1, value="GEOMETRIK DENETIM RAPORU" if lang == "tr" else "GEOMETRIC INSPECTION REPORT").font = Font(size=16, bold=True, color="1F4E79")
    ws1.merge_cells("A1:F1")
    ws1.cell(row=2, column=1, value=f"Meter Type: {report.meter_type}  |  NPS: {report.nps}  |  D: {report.D_mm} mm  |  beta: {report.beta}")
    ws1.cell(row=3, column=1, value=f"Conditioner: {report.conditioner_type or 'None'}")
    ws1.cell(row=5, column=1, value=f"Overall: {report.overall_status}").font = Font(bold=True, size=13, color="1F4E79")
    ws1.cell(row=6, column=1, value=f"Pass Rate: {report.pass_rate:.1f}% ({report.passed_params}/{report.total_params})")

    # Sheet 2: All parameters
    ws2 = wb.create_sheet("Tum Parametreler" if lang == "tr" else "All Parameters")
    headers = ["Component", "Parametre", "Birim", "Kritiklik", "Olculen", "Tolerans", "Durum", "Standart"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers))

    r = 2
    for comp in report.components:
        for param in comp.parameters:
            ws2.cell(row=r, column=1, value=comp.component_name)
            ws2.cell(row=r, column=2, value=param.label)
            ws2.cell(row=r, column=3, value=param.unit)
            ws2.cell(row=r, column=4, value=param.criticality)
            if param.is_qualitative:
                ws2.cell(row=r, column=5, value=param.qualitative_value or "-")
                ws2.cell(row=r, column=6, value="Enum")
            else:
                vals = [str(p.measured) for p in param.points if p.measured is not None]
                ws2.cell(row=r, column=5, value=", ".join(vals) if vals else "-")
                if param.points and param.points[0].tol_upper is not None:
                    ws2.cell(row=r, column=6, value=f"±{param.points[0].tol_upper:.3f}")
                else:
                    ws2.cell(row=r, column=6, value="-")
            ws2.cell(row=r, column=7, value=param.overall_status)
            ws2.cell(row=r, column=8, value=param.standard_clause)
            for c in range(1, len(headers) + 1):
                ws2.cell(row=r, column=c).border = thin_border
            fill = status_fill(param.overall_status)
            if fill:
                ws2.cell(row=r, column=7).fill = fill
            r += 1

    for c in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 22

    # Sheet 3: Failed only
    failed = report.get_failed_parameters()
    if failed:
        ws3 = wb.create_sheet("Hatalar" if lang == "tr" else "Failures")
        headers3 = ["Component", "Parametre", "Kritiklik", "Durum", "Standart", "Olculen", "Aciklama"]
        for c, h in enumerate(headers3, 1):
            ws3.cell(row=1, column=c, value=h)
        style_header(ws3, 1, len(headers3))
        for i, param in enumerate(failed, 2):
            comp = next((c for c in report.components if param in c.parameters), None)
            ws3.cell(row=i, column=1, value=comp.component_name if comp else "")
            ws3.cell(row=i, column=2, value=param.label)
            ws3.cell(row=i, column=3, value=param.criticality)
            ws3.cell(row=i, column=4, value=param.overall_status)
            ws3.cell(row=i, column=5, value=param.standard_clause)
            ws3.cell(row=i, column=7, value="Tolerans disi" if lang == "tr" else "Out of tolerance")
            ws3.cell(row=i, column=4).fill = fail_fill
            for c in range(1, len(headers3) + 1):
                ws3.cell(row=i, column=c).border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
