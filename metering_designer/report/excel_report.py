"""
Excel report generation for metering station design.
"""

import io
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from metering_designer.core.weights import CATEGORY_LABELS_TR


def _fmt_labels(lang: str = "tr") -> dict:
    """Excel labels in TR or EN; chart/sheet titles localize."""
    en = lang != "tr"
    return {
        "title": "ÖLÇÜM İSTASYONU DİZAYN RAPORU" if not en else "METERING STATION DESIGN REPORT",
        "sub1": "BU RAPOR NASIL OKUNMALI?" if not en else "HOW TO READ THIS REPORT?",
        "intro1": ("Bu rapor, olcum istasyonunuz icin en uygun akis metre tipini belirlemek amaciyla hazirlanmistir."
                   if not en else
                   "This report is prepared to determine the most suitable flow meter type for your metering station."),
        "intro2": ("Program 6 kategoride 30+ kriteri degerlendirerek her metre tipine 0-100 arasi bir puan verir. "
                   "En yuksek puan = en uygun secenek."
                   if not en else
                   "The program evaluates 30+ criteria in 6 categories and scores each meter type 0-100. "
                   "Highest score = best fit."),
        "sel_meter": "SECILEN METRE" if not en else "SELECTED METER",
        "type": "Tip:" if not en else "Type:",
        "score": "Puan:" if not en else "Score:",
        "proc_sum": "PROSES ÖZETI" if not en else "PROCESS SUMMARY",
        "process_summary_keys": [("Akışkan", "Q min/nom/max", "P işl / tas", "T işl / tas", "NPS")] if not en
                                else [("Fluid", "Q min/nom/max", "P op / ds", "T op / ds", "NPS")],
        "sheet_scoring": "Puanlama" if not en else "Scoring",
        "scoring_title": "METRE PUANLAMA DETAYI" if not en else "METER SCORING DETAIL",
        "scoring_headers": (["Sıra", "Metre Tipi", "Toplam Puan", "Tier"]
                            if not en else ["No", "Meter Type", "Total Score", "Tier"]),
        "sheet_detail": "Detay Puan" if not en else "Detail Score",
        "detail_title": "DETAY PUANLAMA: {m}" if not en else "DETAIL SCORING: {m}",
        "detail_headers": (["Kategori", "Kriter", "Puan (0-10)", "Ağırlık", "Açıklama"]
                           if not en else ["Category", "Criterion", "Score (0-10)", "Weight", "Description"]),
        "weight_fmt": "{:.2f} (ağırlık: {:.0f}%)" if not en else "{:.2f} (weight: {:.0f}%)",
        "sheet_cond": "Akis Düzenleyici" if not en else "Flow Conditioner",
        "cond_title": "AKIŞ DÜZENLEYICI PUANLAMA" if not en else "FLOW CONDITIONER SCORING",
        "cond_headers": (["Sıra", "Tip", "Toplam Puan", "K Faktörü", "Düz Boru Azaltma", "ISO Uyumlu", "Etkili Uzunluk (m)"]
                         if not en else ["No", "Type", "Total Score", "K Factor", "Straight-Pipe Reduction",
                                         "ISO Compliant", "Effective Length (m)"]),
        "yes": "Evet" if not en else "Yes",
        "no": "Hayır" if not en else "No",
        "sheet_std": "Standartlar" if not en else "Standards",
        "std_title": "STANDARTLAR KONTROL LISTESI" if not en else "STANDARDS CHECKLIST",
        "std_headers": ["Standart", "Uygulandı", "Not"] if not en else ["Standard", "Applied", "Note"],
        "std_notes": {
            "pipe": "Proses borulama dizaynı" if not en else "Process piping design",
            "flange": "Flanş seçimi" if not en else "Flange selection",
            "sour": "Sour servis malzeme" if not en else "Sour service material",
            "ex": "Ex zone sınıflandırma" if not en else "Ex zone classification",
            "sil": "SIL değerlendirme" if not en else "SIL assessment",
            "unc": "Belirsizlik bütçesi" if not en else "Uncertainty budget",
            "hv": "Gaz ısıl değer hesabı" if not en else "Gas heating value",
            "meter_std": "Metre tipi standardı" if not en else "Meter type standard",
        },
    }


def generate_excel_report(
    project: dict,
    process: dict,
    requirements: dict,
    scored_meters: list,
    selected_meter,
    engineering: dict,
    conditioners: Optional[list] = None,
    lang: str = "tr",
) -> io.BytesIO:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl gereklidir: pip install openpyxl")

    wb = Workbook()

    # --- Colors ---
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Ozet"
    L = _fmt_labels(lang)

    ws1.cell(row=1, column=1, value=L["title"]).font = title_font
    ws1.merge_cells("A1:D1")
    ws1.cell(row=2, column=1, value=f"Proje: {project.get('name','-')}  |  Konum: {project.get('location','-')}")
    ws1.cell(row=3, column=1, value=f"Tarih: {project.get('date','-')}  |  Etiket: {project.get('tag','-')}")

    ws1.cell(row=5, column=1, value=L["sub1"]).font = Font(bold=True, size=12, color="C62828")
    ws1.merge_cells("A5:D5")
    ws1.cell(row=6, column=1, value=L["intro1"]).font = Font(italic=True)
    ws1.merge_cells("A6:D6")
    ws1.cell(row=7, column=1, value=L["intro2"])
    ws1.merge_cells("A7:D7")
    ws1.cell(row=9, column=1, value=L["sel_meter"]).font = Font(bold=True, size=12)
    ws1.cell(row=10, column=1, value=f"{L['type']} {selected_meter.name_tr}")
    ws1.cell(row=10, column=2, value=f"{L['score']} {selected_meter.total_score}/100")
    ws1.cell(row=10, column=3, value=f"Tier: {selected_meter.tier_label}")

    ws1.cell(row=12, column=1, value=L["proc_sum"]).font = Font(bold=True, size=12)
    pkeys = L["process_summary_keys"][0]
    row = 13
    for k, v in [(pkeys[0], process.get("fluid_type", "-")),
                  (pkeys[1], f"{process.get('qmin',0)}/{process.get('qnormal',0)}/{process.get('qmax',0)}"),
                  (pkeys[2], f"{process.get('oper_p_bar',0)} / {process.get('design_p_bar',0)} barg"),
                  (pkeys[3], f"{process.get('oper_t_c',0)} / {process.get('design_t_c',0)} °C"),
                  (pkeys[4], process.get("nps", "-"))]:
        ws1.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=str(v))
        row += 1

    # Sheet 2: Scoring Details
    ws2 = wb.create_sheet(L["sheet_scoring"])
    ws2.cell(row=1, column=1, value=L["scoring_title"]).font = title_font
    ws2.merge_cells("A1:G1")

    headers = L["scoring_headers"] + list(CATEGORY_LABELS_TR.values())
    for c, h in enumerate(headers, 1):
        ws2.cell(row=3, column=c, value=h)
    style_header(ws2, 3, len(headers))

    for i, m in enumerate(scored_meters, 1):
        r = i + 3
        ws2.cell(row=r, column=1, value=i)
        ws2.cell(row=r, column=2, value=m.name_tr)
        ws2.cell(row=r, column=3, value=m.total_score)
        ws2.cell(row=r, column=4, value=m.tier_label)
        col = 5
        for ck in CATEGORY_LABELS_TR:
            cat = m.categories.get(ck)
            ws2.cell(row=r, column=col, value=round(cat.score, 2) if cat else 0)
            col += 1

    for c in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    # Sheet 3: Detailed Scoring (selected meter)
    ws3 = wb.create_sheet(L["sheet_detail"])
    ws3.cell(row=1, column=1, value=L["detail_title"].format(m=selected_meter.name_tr)).font = title_font
    ws3.cell(row=3, column=1, value=L["detail_headers"][0]).font = Font(bold=True)
    ws3.cell(row=3, column=2, value=L["detail_headers"][1]).font = Font(bold=True)
    ws3.cell(row=3, column=3, value=L["detail_headers"][2]).font = Font(bold=True)
    ws3.cell(row=3, column=4, value=L["detail_headers"][3]).font = Font(bold=True)
    ws3.cell(row=3, column=5, value=L["detail_headers"][4]).font = Font(bold=True)
    style_header(ws3, 3, 5)

    r = 4
    for ck, cl in CATEGORY_LABELS_TR.items():
        cat = selected_meter.categories.get(ck)
        if not cat:
            continue
        ws3.cell(row=r, column=1, value=cl).font = Font(bold=True)
        ws3.cell(row=r, column=2, value=L["weight_fmt"].format(cat.score, cat.weight * 100))
        r += 1
        for crit in cat.criteria:
            ws3.cell(row=r, column=2, value=crit.name)
            ws3.cell(row=r, column=3, value=round(crit.score, 2))
            ws3.cell(row=r, column=4, value=round(crit.weight * 100, 0))
            ws3.cell(row=r, column=5, value=crit.justification)
            r += 1

    ws3.column_dimensions["E"].width = 60

    # Sheet 4: Flow Conditioner
    if conditioners:
        ws4 = wb.create_sheet(L["sheet_cond"])
        ws4.cell(row=1, column=1, value=L["cond_title"]).font = title_font
        headers_c = L["cond_headers"]
        for c, h in enumerate(headers_c, 1):
            ws4.cell(row=3, column=c, value=h)
        style_header(ws4, 3, len(headers_c))
        for i, cond in enumerate(conditioners[:5], 1):
            r = i + 3
            ws4.cell(row=r, column=1, value=i)
            ws4.cell(row=r, column=2, value=cond["name_tr"])
            ws4.cell(row=r, column=3, value=cond["total_score"])
            ws4.cell(row=r, column=4, value=cond["k_factor"])
            ws4.cell(row=r, column=5, value=cond["reduction_pct"])
            ws4.cell(row=r, column=6, value=L["yes"] if cond["iso_compliant"] else L["no"])
            ws4.cell(row=r, column=7, value=cond["effective_length_m"])

    # Sheet 5: Standards Checklist
    ws5 = wb.create_sheet(L["sheet_std"])
    ws5.cell(row=1, column=1, value=L["std_title"]).font = title_font
    ws5.cell(row=3, column=1, value=L["std_headers"][0]).font = Font(bold=True)
    ws5.cell(row=3, column=2, value=L["std_headers"][1]).font = Font(bold=True)
    ws5.cell(row=3, column=3, value=L["std_headers"][2]).font = Font(bold=True)
    style_header(ws5, 3, 3)

    standards = [
        ("ASME B31.3", "✓", L["std_notes"]["pipe"]),
        ("ASME B16.5", "✓", L["std_notes"]["flange"]),
        ("ISO 15156 / NACE MR0175", "✓" if requirements.get("h2s") else "—", L["std_notes"]["sour"]),
        ("IEC 60079-10-1", "✓", L["std_notes"]["ex"]),
        ("IEC 61511", "✓", L["std_notes"]["sil"]),
        ("ISO 5168", "✓", L["std_notes"]["unc"]),
        ("ISO 6976", "✓", L["std_notes"]["hv"]),
    ]
    meter_standards = {"ultrasonic": "AGA 9 / ISO 17089", "orifice": "ISO 5167-2 / AGA 3",
                        "turbine": "AGA 7 / ISO 9951", "coriolis": "AGA 11 / ISO 10790",
                        "vortex": "ISO 17089-2", "positive_displacement": "API MPMS Ch.4"}
    if selected_meter.meter_key in meter_standards:
        standards.insert(0, (meter_standards[selected_meter.meter_key], "✓", L["std_notes"]["meter_std"]))

    for i, (std, status, note) in enumerate(standards, 4):
        ws5.cell(row=i, column=1, value=std)
        ws5.cell(row=i, column=2, value=status)
        ws5.cell(row=i, column=3, value=note)
    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["C"].width = 40

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
