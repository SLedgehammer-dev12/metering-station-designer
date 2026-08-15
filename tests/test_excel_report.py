"""Excel report generation tests."""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from collections import namedtuple
from metering_designer.report.excel_report import generate_excel_report, HAS_OPENPYXL
from metering_designer.core.weights import CATEGORY_LABELS_TR


# -- Mock helpers following test_phase3.py pattern --
SMock = namedtuple('SMock', ['meter_key','name_tr','name_en','total_score',
                              'tier_label','tier_color','categories',
                              'strengths','weaknesses','details'])
CatMock = namedtuple('CatMock', ['score','weight','criteria'])
CritMock = namedtuple('CritMock', ['name','score','weight','justification'])


def _build_mock_selected_meter():
    """Return a selected meter mock with all categories populated."""
    cats = {}
    for ck in CATEGORY_LABELS_TR:
        cats[ck] = CatMock(8.0, 0.15, [
            CritMock('Test Kriteri 1', 8.5, 0.4, 'Uygun'),
            CritMock('Test Kriteri 2', 7.0, 0.6, 'Kismen uygun'),
        ])
    return SMock('ultrasonic', 'Ultrasonik (USM)', 'Ultrasonic (USM)',
                 78.5, '\u2605\u2605\u2606', 'blue', cats,
                 ['Guclu yon 1'], ['Zayif yon 1'], {})


def _build_mock_scored_meters():
    """Return a list of 3 scored meter mocks."""
    meters = []
    for idx, (key, name_tr, score, tier) in enumerate([
        ('ultrasonic', 'Ultrasonik (USM)', 78.5, '\u2605\u2605\u2606'),
        ('orifice', 'Orifis Plakasi', 65.2, '\u2605\u2605'),
        ('turbine', 'Turbo Metre', 52.0, '\u2605'),
    ]):
        cats = {}
        for ck in CATEGORY_LABELS_TR:
            cats[ck] = CatMock(7.0 + idx, 0.15, [])
        meters.append(SMock(key, name_tr, name_tr, score, tier, 'blue',
                            cats, [], [], {}))
    return meters


def test_excel_report_generates():
    """Verify generate_excel_report returns non-empty BytesIO buffer."""
    if not HAS_OPENPYXL:
        return  # skip if openpyxl not installed

    project = {'name': 'Test Istasyonu', 'location': 'Marmara',
               'date': '2025-01-01', 'tag': 'FE-101'}
    process = {'fluid_type': 'dogal_gaz', 'qmin': 1000, 'qnormal': 5000,
               'qmax': 10000, 'oper_p_bar': 40, 'design_p_bar': 50,
               'oper_t_c': 30, 'design_t_c': 60, 'nps': '8 inch'}
    requirements = {'h2s': False}
    scored_meters = _build_mock_scored_meters()
    selected_meter = _build_mock_selected_meter()
    engineering = {}

    buf = generate_excel_report(project, process, requirements,
                                scored_meters, selected_meter,
                                engineering)

    assert buf is not None
    data = buf.getvalue()
    assert isinstance(data, bytes)
    assert len(data) > 0, 'Excel buffer should not be empty'


def test_excel_report_buffer_size():
    """Verify generated Excel buffer exceeds 1000 bytes."""
    if not HAS_OPENPYXL:
        return

    project = {'name': 'T', 'location': 'L', 'date': '2025-01-01', 'tag': 'FE-102'}
    process = {'fluid_type': 'dogal_gaz', 'qmin': 1000, 'qnormal': 5000,
               'qmax': 10000, 'oper_p_bar': 40, 'design_p_bar': 50,
               'oper_t_c': 30, 'design_t_c': 60, 'nps': '8 inch'}
    requirements = {'h2s': False}
    scored_meters = _build_mock_scored_meters()
    selected_meter = _build_mock_selected_meter()
    engineering = {}

    buf = generate_excel_report(project, process, requirements,
                                scored_meters, selected_meter,
                                engineering)
    assert len(buf.getvalue()) > 1000, (
        f'Expected >1000 bytes, got {len(buf.getvalue())}'
    )


def test_excel_report_sheets_exist():
    """Verify expected sheet names exist in the workbook."""
    if not HAS_OPENPYXL:
        return

    from openpyxl import load_workbook

    project = {'name': 'T', 'location': 'L', 'date': '2025-01-01', 'tag': 'FE-103'}
    process = {'fluid_type': 'dogal_gaz', 'qmin': 1000, 'qnormal': 5000,
               'qmax': 10000, 'oper_p_bar': 40, 'design_p_bar': 50,
               'oper_t_c': 30, 'design_t_c': 60, 'nps': '8 inch'}
    requirements = {'h2s': False}
    scored_meters = _build_mock_scored_meters()
    selected_meter = _build_mock_selected_meter()
    engineering = {}

    # First test without conditioners (4 sheets)
    buf_no_cond = generate_excel_report(project, process, requirements,
                                        scored_meters, selected_meter,
                                        engineering)
    wb = load_workbook(buf_no_cond)
    sheet_names = wb.sheetnames

    expected_sheets = ['Ozet', 'Puanlama', 'Detay Puan', 'Standartlar']
    for sname in expected_sheets:
        assert sname in sheet_names, f'Sheet "{sname}" not found in {sheet_names}'

    # Test with conditioners (5 sheets, includes "Akis Düzenleyici")
    conditioners = [{
        'name_tr': 'Zanker Plakasi',
        'total_score': 85.0,
        'k_factor': 0.6,
        'reduction_pct': 50,
        'iso_compliant': True,
        'effective_length_m': 2.5,
    }]
    buf_with_cond = generate_excel_report(project, process, requirements,
                                          scored_meters, selected_meter,
                                          engineering, conditioners)
    wb2 = load_workbook(buf_with_cond)
    sheet_names2 = wb2.sheetnames

    assert 'Akis Düzenleyici' in sheet_names2, (
        f'"Akis Düzenleyici" should exist when conditioners provided: {sheet_names2}'
    )


def test_excel_report_sheets_count():
    """Verify correct number of sheets with and without conditioners."""
    if not HAS_OPENPYXL:
        return

    from openpyxl import load_workbook

    project = {'name': 'T', 'location': 'L', 'date': '2025-01-01', 'tag': 'FE-104'}
    process = {'fluid_type': 'dogal_gaz', 'qmin': 1000, 'qnormal': 5000,
               'qmax': 10000, 'oper_p_bar': 40, 'design_p_bar': 50,
               'oper_t_c': 30, 'design_t_c': 60, 'nps': '8 inch'}
    requirements = {'h2s': True}
    scored_meters = _build_mock_scored_meters()
    selected_meter = _build_mock_selected_meter()
    engineering = {}

    # Without conditioners: Ozet, Puanlama, Detay Puan, Standartlar = 4 sheets
    buf = generate_excel_report(project, process, requirements,
                                scored_meters, selected_meter,
                                engineering)
    wb = load_workbook(buf)
    assert len(wb.sheetnames) == 4, (
        f'Expected 4 sheets without conditioners, got {len(wb.sheetnames)}: {wb.sheetnames}'
    )

    # With conditioners: 5 sheets (adds "Akis Düzenleyici")
    conditioners = [{
        'name_tr': 'Zanker Plakasi',
        'total_score': 85.0,
        'k_factor': 0.6,
        'reduction_pct': 50,
        'iso_compliant': True,
        'effective_length_m': 2.5,
    }]
    buf2 = generate_excel_report(project, process, requirements,
                                 scored_meters, selected_meter,
                                 engineering, conditioners)
    wb2 = load_workbook(buf2)
    assert len(wb2.sheetnames) == 5, (
        f'Expected 5 sheets with conditioners, got {len(wb2.sheetnames)}: {wb2.sheetnames}'
    )


if __name__ == "__main__":
    for name, func in list(locals().items()):
        if name.startswith("test_"):
            try:
                func()
                print(f"\u2705 {name}")
            except Exception as e:
                print(f"\u274c {name}: {e}")
    print("\nExcel report tests complete")
